import openpyxl
from openpyxl.styles import Font

from metaerg import context
from metaerg.datatypes import sqlite


def write_genomes_to_xls(db_connection):
    excel_file = context.BASE_DIR / 'genome_properties.xls'
    e_workbook = openpyxl.Workbook()
    e_column = 3
    black = Font(color="000000")
    grey = Font(color="AAAAAA")
    green = Font(color="009900")
    rows_for_taxon = max(len(g.top_taxon.split('; ')) for g in sqlite.read_all_genomes(db_connection))
    for genome in sqlite.read_all_genomes(db_connection):
        pretty_properties = genome.to_dict_pretty()
        e_sheet = e_workbook.active
        row = 1
        e_sheet.cell(row=row, column=1).value = 'genome properties'
        for k, v in pretty_properties.items():
            if 'top taxon' == k:
                continue
            e_sheet.cell(row=row, column=2).value = k
            e_sheet.cell(row=row, column=e_column).value = v
            row += 1

        e_sheet.cell(row=row, column=1).value = 'classification (top taxon)'
        for k, w in zip('kingdom phylum class order family genus species rank rank rank rank'.split(),
                        genome.top_taxon.split('; ')):
            e_sheet.cell(row=row, column=2).value = k
            e_sheet.cell(row=row, column=e_column).value = w
            row += 1
        row += max(0, rows_for_taxon - len(genome.top_taxon.split('; ')))

        e_sheet.cell(row=row, column=1).value = 'physiology summary'
        for subsystem, value in genome.subsystem_summary.items():
            if isinstance(value, float):
                value *= 100
                value = round(value, 0)
                if value >= 80:
                    color = green
                elif value < 51:
                    color = grey
                else:
                    color = black
                subsystem += ' (% complete)'
            else:
                if value > 0:
                    color = green
                else:
                    color = grey
                subsystem += ' (# genes)'
            e_sheet.cell(row=row, column=2).value = subsystem
            if color is not grey:
                e_sheet.cell(row=row, column=e_column).value = value
            e_sheet.cell(row=row, column=e_column).font = color
            row += 1

        for subsystem, genes in genome.subsystems.items():
            e_sheet.cell(row=row, column=1).value = subsystem
            for gene, feature_ids in genes.items():
                e_sheet.cell(row=row, column=2).value = gene
                e_sheet.cell(row=row, column=e_column).value = ', '.join(feature_ids)
                row += 1
        e_column += 1
    e_workbook.save(excel_file)